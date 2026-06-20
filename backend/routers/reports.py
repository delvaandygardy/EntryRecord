from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from backend.deps import get_db, get_current_user, require_role
from backend.schemas import ReportRequest
import psycopg2.extras
from datetime import datetime
import io

router = APIRouter(prefix="/api/reports", tags=["Rapports"])


def _fetch(conn, table, date_debut=None, date_fin=None):
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    where = []
    params = []
    if date_debut:
        where.append("timestamp >= %s"); params.append(date_debut)
    if date_fin:
        where.append("timestamp <= %s"); params.append(date_fin + " 23:59:59")
    sql = f"SELECT * FROM {table}"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY timestamp DESC"
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    for r in rows:
        for k, v in r.items():
            if isinstance(v, datetime):
                r[k] = v.strftime("%Y-%m-%d %H:%M:%S")
    return rows


@router.get("/stats")
def stats(conn=Depends(get_db), _=Depends(get_current_user)):
    cur = conn.cursor()
    result = {}
    for t in ("vehicules", "conducteurs", "pietons", "employes", "alertes", "presences"):
        cur.execute(f"SELECT COUNT(*) FROM {t}")
        result[f"{t}_total"] = cur.fetchone()[0]
        if t != "employes":
            cur.execute(f"SELECT COUNT(*) FROM {t} WHERE timestamp::date=CURRENT_DATE")
            result[f"{t}_today"] = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM alertes WHERE traitee=FALSE")
    result["alertes_actives"] = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM employes WHERE statut='Actif'")
    result["employes_actifs"] = cur.fetchone()[0]
    return result


@router.post("/pdf")
def generate_pdf(body: ReportRequest, conn=Depends(get_db),
                 _=Depends(require_role("admin", "superviseur", "operateur"))):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer, PageBreak)
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "reportlab non installé")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    story = []

    today = datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph(f"Rapport — Système d'Enregistrement", styles["Title"]))
    story.append(Paragraph(f"Généré le {today}", styles["Normal"]))
    story.append(Spacer(1, 12))

    table_labels = {
        "vehicules":   ("Véhicules",   ["id","plaque","confidence","point_entree","timestamp"]),
        "conducteurs": ("Conducteurs", ["id","nom","prenom","numero_document","type_document","timestamp"]),
        "pietons":     ("Piétons",     ["id","nom","prenom","numero_document","nationalite","timestamp"]),
        "employes":    ("Employés",    ["id","matricule","nom","prenom","poste","departement","statut"]),
    }

    for tbl in body.tables:
        if tbl not in table_labels:
            continue
        label, cols = table_labels[tbl]
        rows = _fetch(conn, tbl, body.date_debut, body.date_fin)
        story.append(Paragraph(f"{label} ({len(rows)})", styles["Heading2"]))
        data = [cols] + [[str(r.get(c, "")) for c in cols] for r in rows[:500]]
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#1a73e8")),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 7),
            ("GRID",        (0,0), (-1,-1), 0.25, colors.grey),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#f1f3f4")]),
        ]))
        story.append(t)
        story.append(Spacer(1,16))
        if tbl != body.tables[-1]:
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    filename = f"rapport_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                              headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/excel")
def generate_excel(body: ReportRequest, conn=Depends(get_db),
                   _=Depends(require_role("admin", "superviseur", "operateur"))):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "openpyxl non installé")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(bold=True, color="FFFFFF")

    table_labels = {
        "vehicules":   ("Véhicules",   ["id","plaque","confidence","point_entree","timestamp","notes"]),
        "conducteurs": ("Conducteurs", ["id","nom","prenom","numero_document","type_document","date_naissance","timestamp"]),
        "pietons":     ("Piétons",     ["id","nom","prenom","numero_document","type_document","nationalite","timestamp"]),
        "employes":    ("Employés",    ["id","matricule","nom","prenom","poste","departement","telephone","statut"]),
    }

    for tbl in body.tables:
        if tbl not in table_labels:
            continue
        label, cols = table_labels[tbl]
        ws = wb.create_sheet(label)
        rows = _fetch(conn, tbl, body.date_debut, body.date_fin)

        for ci, col in enumerate(cols, 1):
            cell = ws.cell(1, ci, col.upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(rows, 2):
            for ci, col in enumerate(cols, 1):
                ws.cell(ri, ci, str(row.get(col, "") or ""))

        for col_cells in ws.columns:
            ws.column_dimensions[col_cells[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"rapport_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})
